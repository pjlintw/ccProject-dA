# necessary imports
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import bs4 as bs
from selenium import webdriver
import pandas as pd
import os
import os.path
import time

class NBA_STATS_API(object):

	NBA_URL = 'https://stats.nba.com/teams/advanced/?sort=W&dir=-1&Season={}&SeasonType={}'

	'''docstring fro NBA_STATS_API'''
	def __init__(self):
		self.start = time.time()
		parser = argparse.ArgumentParser(
			description='''
			An nba.stats.com API, which allows user access nba official website data easily\n
			Input: list of season and type\n
			Output: .xlsx / .csv			
			'''
		)
		parser.add_argument('-r', help='read file')
		parser.add_argument('-w', help='write .xlsx')
		parser.add_argument('-t', help='season type')
		args = parser.parse_args()

		if args == None:
			print('args==None')

		# fetch command
		if args.r:
			query_lst = self.file2lst(args.r)

		# fetch command
		if args.w:
			write_path = args.w
			write_path = self.parser_write_path(write_path)

		if args.t:
			season_type = self.parser_season_type(args.t)

		# check
		if not query_lst:
			parser.print_help()
		elif not write_path:
			parser.print_help()
		else:
			print('\nCrawling:{} {}\n'.format(season_type, '  '.join(query_lst)))
			print(self.parser_data(query_lst, season_type,write_path))


	def parser_data(self, query_lst, season_type, write_path, timeout=3):

		url = self.NBA_URL

		df_set = list()

		# parser
		for season in query_lst:
			print('\nProcessing: {} {}...'.format(season, season_type))
			browser = webdriver.Chrome('./chromedriver')
			browser.get(url.format(season, season_type))
			page_source = browser.page_source
			browser.close()

			soup = bs.BeautifulSoup(page_source, 'lxml-xml')
			t = soup.find('div', class_='nba-stat-table__overflow').table
			t = str(t)
			df = pd.read_html(t)[0]

			# store excel
			self.store_excel(df, season, write_path)

		# finish msg
		end = time.time()
		msg = '\n{} in path: {}'.format(write_path, os.getcwd().replace('\\','/') + '/' + write_path)
		msg = '\nfinished in {} sec\n'.format(str(end-self.start)) + msg
		return msg


	@staticmethod
	def store_csv(df_set, sheet_name, file_name):
		writer = pd.ExcelWriter(file_name + '.xlsx', engine='xlsxwriter')
		for df in df_set:
			df.to_excel(writer, sheet_name=sheet_name)
		writer.save()

	@staticmethod
	def file2lst(file_path):
		file_type = file_path.split('.')[-1]
		if file_type != 'txt':
			file_path = file_path + '.txt'
		else:
			file_path = file_path + '.txt'

		file_path = os.getcwd() + '/read_zone' + '/' + file_path
		f = open(file_path, 'r', encoding='utf-8').read()
		return [i for i in f.split('\n') if i != '']

	@staticmethod
	def parser_write_path(write_path):
		if '.xlsx' not in write_path and '.csv' not in write_path:
			return write_path + '.xlsx'
		elif '.csv' in write_path:
			return ''.join(write_path.split('.')[:-1]) + '.xlsx'
		else:
			return write_path


	@staticmethod
	def parser_season_type(raw_season_type):
		raw_season_type = raw_season_type.lower()
		if raw_season_type == 'preseason' or raw_season_type == 'pre season' or raw_season_type == 'pre-season':
			return 'Pre%20Season'

		elif raw_season_type == 'regular season' or raw_season_type == 'regular-season':
			return 'Regular%20Season'

		elif raw_season_type == 'playoffs' or raw_season_type == 'playoff':
			return 'Playoffs'

		elif raw_season_type == 'all star' or raw_season_type == 'all-star' or raw_season_type == 'allstar':
			return 'All%20Star'

	@staticmethod
	def html_table2df(soup):
		t = str(soup.find('div', class_='nba-stat-table__overflow')['table'])
		df = pd.read_html(t)[0]
		df.rename(columns={'Unnamed: 0': 'rank'}, inplace=True)
		df = df.iloc[:, :df.columns.get_loc('PIE') + 1]
		return df

	@staticmethod
	def store_excel(df, sheet_name, file_name):
		df.rename(columns={'Unnamed: 0': 'rank'}, inplace=True)
		dataframe = df.iloc[:, :df.columns.get_loc('PIE') + 1]

		if not os.path.isfile(file_name):
			# create new and save it
			# if file existed, new one will overwrite !!!!!
			workbook_name = file_name
			wb = Workbook()
			wb.save(workbook_name)

		wb = load_workbook(file_name)
		wb.create_sheet(sheet_name)
		ws = wb.get_sheet_by_name(sheet_name)
		for r in dataframe_to_rows(dataframe, index=False, header=True):
			ws.append(r)
		wb.save(file_name)


if __name__ == '__main__':
	c = NBA_STATS_API()